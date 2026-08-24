"""Writing a renderer-neutral document as a workbook.

Two sheets, because they answer different questions. `Records` is the table
alone, so its first row is a header a spreadsheet can turn into a filter.
`Details` is where the file says what it is -- the header text, a record
report's own fields, each group's values and subtotal, and the grand total.

One rule for every document that reaches here, browse export or report. A
grouped listing flattens the way `render_csv` flattens it, with the group
values repeated as leading columns, because a spreadsheet pivots for itself
and a banded sheet is not a table. The subtotals still travel, on sheet two,
because they are not always recomputable: `avg` is the report service dividing
HALF_EVEN at field scale, which `AVERAGE()` will not reproduce.

Numbers arrive as numbers. That is the only reason to prefer this over CSV.

It is *not* automatically safer than CSV about formulas, which was the
assumption this module started from and a test disproved: openpyxl infers a
cell's type from what it is handed, and infers `formula` for anything starting
with `=`. A workbook is safe once that guess is overruled, and better than
CSV's apostrophe when it is, because the value survives intact rather than
carrying an escape into whatever reads it back.

Optional, like PDF. A server without the `spreadsheet` extra says so rather
than failing on an import nobody asked for, and the presentation manifest does
not offer the format at all -- so a renderer can never present a download the
server would refuse.
"""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Sequence

from .document import ReportDocument

try:  # pragma: no cover - the absent branch runs only without the extra
    from openpyxl import Workbook

    SPREADSHEET_AVAILABLE = True
except ModuleNotFoundError:  # pragma: no cover
    Workbook = None  # type: ignore[assignment,misc]
    SPREADSHEET_AVAILABLE = False

RECORDS_SHEET = "Records"
DETAILS_SHEET = "Details"
REPORT_TOTAL_HEADING = "Report total"

TypedValues = Sequence[Sequence[Any]]
"""Typed cells positional over `document.detail.rows`, or empty for none.

Positional over the *detail* rows rather than over what the sheet shows: a
grouped listing gains leading group columns, and a group key is a caption
rather than a value, so it is always text.
"""


def render_xlsx(
    document: ReportDocument,
    typed_values: TypedValues = (),
) -> bytes:
    """Render one document as workbook bytes."""

    if not SPREADSHEET_AVAILABLE:
        raise RuntimeError(
            "XLSX export needs the 'spreadsheet' extra: "
            "pip install tide-framework[spreadsheet]"
        )
    book = Workbook()
    _write_records(book.active, document, typed_values)
    _write_details(book.create_sheet(DETAILS_SHEET), document)
    buffer = BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def write_xlsx(
    document: ReportDocument,
    path: str | Path,
    typed_values: TypedValues = (),
) -> Path:
    """Write the workbook to a file and return its path."""

    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_bytes(render_xlsx(document, typed_values))
    return output


def _write_records(
    sheet: Any,
    document: ReportDocument,
    typed_values: TypedValues,
) -> None:
    """Sheet one: the flat table, and nothing that is not table."""

    sheet.title = RECORDS_SHEET
    leading = tuple(
        value.label for value in document.groups[0].values
    ) if document.groups else ()
    for position, label in enumerate(leading, start=1):
        _text(sheet.cell(row=1, column=position), label)
    for offset, column in enumerate(document.detail.columns, start=1):
        _text(sheet.cell(row=1, column=len(leading) + offset), column.label)

    # A group's values repeat on every row of its slice, because a spreadsheet
    # has no headings to put them in and a row that cannot say whose it is
    # cannot be sorted, filtered or pivoted.
    prefixes: list[tuple[str, ...]] = [() for _ in document.detail.rows]
    for group in document.groups:
        for index in range(group.row_start, group.row_start + group.row_count):
            if index < len(prefixes):
                prefixes[index] = tuple(value.text for value in group.values)

    for index, row in enumerate(document.detail.rows):
        typed = typed_values[index] if index < len(typed_values) else ()
        for position, text in enumerate(prefixes[index], start=1):
            _text(sheet.cell(row=index + 2, column=position), text)
        for offset, cell in enumerate(row, start=1):
            target = sheet.cell(row=index + 2, column=len(leading) + offset)
            value = typed[offset - 1] if offset - 1 < len(typed) else None
            if value is None:
                _text(target, cell.text)
            elif isinstance(value, datetime):
                # A workbook has no timezone to keep, and openpyxl refuses an
                # aware value rather than quietly dropping the offset.
                target.value = value.replace(tzinfo=None)
            else:
                target.value = value


def _write_details(sheet: Any, document: ReportDocument) -> None:
    """Sheet two: everything a table has no room for."""

    line = 1
    for text in document.header_text:
        _text(sheet.cell(row=line, column=1), text)
        line += 1
    line = _write_values(sheet, line, document.record_values)
    for group in document.groups:
        line += 1 if line > 1 else 0
        line = _write_values(sheet, line, group.values)
        line = _write_values(sheet, line, group.footer_values)
    if document.footer_values:
        line += 1 if line > 1 else 0
        if document.groups:
            # A group's subtotal and the report's grand total are the same
            # labels over different scopes, so without this the sheet shows
            # the same block twice and says nothing about which is which.
            # The PDF's bands make that obvious; a list of pairs does not.
            _text(sheet.cell(row=line, column=1), REPORT_TOTAL_HEADING)
            line += 1
        _write_values(sheet, line, document.footer_values)


def _write_values(sheet: Any, line: int, values: Any) -> int:
    for value in values:
        _text(sheet.cell(row=line, column=1), value.label)
        _text(sheet.cell(row=line, column=2), value.text)
        line += 1
    return line


def _text(cell: Any, value: str) -> None:
    """Write one string, as a string, whatever it starts with.

    openpyxl infers a cell's type from what it is given, and infers `formula`
    for anything starting with `=`. So a workbook is *not* automatically safe
    from application data that looks like a formula -- the guess has to be
    overruled, or a stored `=SUM(A1:A9)` becomes something Excel evaluates.

    Overruling it rather than escaping it is what keeps the value intact: the
    apostrophe CSV needs would still be there when the cell was read back.
    """

    cell.value = value
    cell.data_type = "s"
