"""What a browse export looks like once it is a file.

CSV is the table and nothing else -- its whole purpose is to be sorted,
filtered and pivoted, and a preamble row breaks exactly that. The workbook has
room, so it carries the provenance on a second sheet and holds numbers as
numbers, which is the only reason to prefer it.
"""

from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Iterator

import pytest

from tide import compile_project
from tide.data import InMemoryRepository, SortField
from tide.reporting.browse import BrowseExport, BrowseExportService
from tide.reporting.csv import render_csv
from tide.reporting.xlsx import SPREADSHEET_AVAILABLE, render_xlsx
from tide.runtime import Channel, Principal, RequestContext
from tide.services import RecordsService

ROOT = Path(__file__).parents[1]
INVOICING = ROOT / "applications" / "invoicing"

VIEW = "sales.Invoice.browse"

CUSTOMERS = [
    {
        "id": 1,
        "code": "ACME",
        "name": "ACME Ltd",
        "email": None,
        "active": True,
        "invoices": [],
    },
]

INVOICES = [
    {
        "id": 1,
        "number": "INV-2026-0001",
        "invoice_date": date(2026, 7, 1),
        "currency": "EUR",
        "status": "draft",
        "posted_at": None,
        "posted_by": None,
        "version": 1,
        "customer": 1,
        "total": Decimal("10.50"),
        "lines": [],
    },
    {
        # A legacy database holds whatever was written into it, and TIDE does
        # not validate rows it did not write.
        "id": 2,
        "number": "=SUM(A1:A9)",
        "invoice_date": date(2026, 7, 2),
        "currency": "EUR",
        "status": "draft",
        "posted_at": None,
        "posted_by": None,
        "version": 1,
        "customer": 1,
        "total": Decimal("20.50"),
        "lines": [],
    },
]


def _context() -> RequestContext:
    return RequestContext(
        principal=Principal("user:clerk", roles=frozenset({"sales_clerk"})),
        channel=Channel.REST,
    )


@pytest.fixture
def export() -> Iterator[BrowseExport]:
    model = compile_project(INVOICING)
    repository = InMemoryRepository()
    repository.seed("crm.Customer", CUSTOMERS)
    repository.seed("sales.Invoice", INVOICES)
    service = BrowseExportService(model, RecordsService(model, repository))
    yield service.build(VIEW, (), (SortField("id"),), _context())


def test_csv_is_the_table_and_only_the_table(export: BrowseExport) -> None:
    text = render_csv(export.document)
    lines = text.strip().splitlines()

    assert len(lines) == 1 + export.rows
    assert lines[0].startswith("Number,")
    # The provenance is real, and deliberately not in the file: a preamble
    # would break the one thing a CSV is for.
    assert export.document.header_text
    assert export.document.application not in text
    # A formula-shaped value is still defused for the spreadsheet that opens
    # it, because a CSV cell has no type to protect it.
    assert "'=SUM(A1:A9)" in text


@pytest.mark.skipif(not SPREADSHEET_AVAILABLE, reason="spreadsheet extra absent")
def test_the_workbook_holds_numbers_as_numbers_and_says_where_it_came_from(
    export: BrowseExport,
) -> None:
    from openpyxl import load_workbook

    book = load_workbook(BytesIO(render_xlsx(export)))

    assert book.sheetnames == ["Records", "Export details"]
    sheet = book["Records"]
    headers = [cell.value for cell in sheet[1]]
    assert headers == ["Number", "Invoice Date", "Customer", "Status", "Total"]

    # A number, not "10.50" -- the whole reason to prefer a workbook.
    total_at = headers.index("Total") + 1
    assert sheet.cell(row=2, column=total_at).value == Decimal("10.50")
    assert sheet.cell(row=2, column=total_at).data_type == "n"

    date_at = headers.index("Invoice Date") + 1
    assert sheet.cell(row=2, column=date_at).value.date() == date(2026, 7, 1)

    # A reference keeps the name the reader saw, never its stored identity.
    customer_at = headers.index("Customer") + 1
    assert sheet.cell(row=2, column=customer_at).value == "ACME - ACME Ltd"

    # A typed string cell is never a formula, so the CSV guard is unnecessary
    # here -- and would have corrupted the value if it had been applied.
    number_at = headers.index("Number") + 1
    assert sheet.cell(row=3, column=number_at).value == "=SUM(A1:A9)"
    assert sheet.cell(row=3, column=number_at).data_type == "s"

    details = [row[0].value for row in book["Export details"].iter_rows()]
    assert details == list(export.document.header_text)


@pytest.mark.skipif(SPREADSHEET_AVAILABLE, reason="spreadsheet extra present")
def test_without_the_extra_the_workbook_refuses_by_name(
    export: BrowseExport,
) -> None:
    with pytest.raises(RuntimeError, match="spreadsheet"):
        render_xlsx(export)
