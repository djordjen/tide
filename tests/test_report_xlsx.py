"""A report as a workbook.

The same rule browse export settled on: sheet one is the flat table and
nothing else, so its first row is a header Excel can filter and pivot on;
sheet two carries everything that is not table -- the header text, a record
report's own fields, each group's values and subtotal, and the grand total.

A grouped listing flattens the way the CSV writer flattens it, with the group
values repeated as leading columns, because a spreadsheet pivots for itself
and a banded sheet is not a table. Its subtotals still travel, on sheet two,
because they are not always recomputable: `avg` is the service dividing
HALF_EVEN at field scale, which `AVERAGE()` will not reproduce.
"""

from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from pathlib import Path

import pytest

from tide import compile_project
from tide.data import InMemoryRepository
from tide.reporting import (
    ReportCell,
    ReportColumn,
    ReportDocument,
    ReportGroup,
    ReportService,
    ReportTable,
    ReportValue,
    render_csv,
)
from tide.reporting.xlsx import (
    DETAILS_SHEET,
    RECORDS_SHEET,
    SPREADSHEET_AVAILABLE,
    render_xlsx,
)
from tide.runtime import Channel, Principal, RequestContext
from tide.services import RecordsService
from tide.tui import seed_demo_data

ROOT = Path(__file__).parents[1]
INVOICING = ROOT / "applications" / "invoicing"

pytestmark = pytest.mark.skipif(
    not SPREADSHEET_AVAILABLE, reason="spreadsheet extra absent"
)


@pytest.fixture
def reporting() -> tuple[ReportService, RequestContext]:
    model = compile_project(INVOICING)
    repository = InMemoryRepository()
    assert seed_demo_data(model, repository) == 15
    service = ReportService(model, RecordsService(model, repository))
    context = RequestContext(
        Principal("report:user", roles=frozenset({"sales_clerk"})),
        channel=Channel.TUI,
    )
    return service, context


def _sheets(document: ReportDocument, typed: tuple[tuple[object, ...], ...] = ()):
    from openpyxl import load_workbook

    return load_workbook(BytesIO(render_xlsx(document, typed)))


def test_a_grouped_listing_flattens_exactly_as_the_csv_does(
    reporting: tuple[ReportService, RequestContext],
) -> None:
    """One rule for the rows, so the two exports cannot disagree."""

    service, context = reporting
    document = service.build("sales.summary", {}, context)
    assert document.groups, "the fixture must be a grouped listing"

    book = _sheets(document)
    sheet = book[RECORDS_SHEET]
    rows = [[cell.value for cell in row] for row in sheet.iter_rows()]

    csv_rows = [
        line.split(",") for line in render_csv(document).strip().splitlines()
    ]
    assert len(rows) == len(csv_rows)
    # The group values lead every row, so a row still says whose it is.
    assert rows[0][:2] == ["Customer", "Currency"]
    assert rows[0] == [
        "Customer",
        "Currency",
        "Number",
        "Invoice Date",
        "Total",
    ]
    assert all(row[0] and row[1] for row in rows[1:])


def test_sheet_two_carries_what_a_table_has_no_room_for(
    reporting: tuple[ReportService, RequestContext],
) -> None:
    service, context = reporting
    document = service.build("sales.summary", {}, context)

    book = _sheets(document)
    lines = [
        [cell.value for cell in row]
        for row in book[DETAILS_SHEET].iter_rows()
    ]
    flat = [
        " ".join(str(value) for value in line if value is not None)
        for line in lines
    ]

    # Each group names itself and closes with its own subtotal.
    assert any("Customer" in text and "Adria" in text for text in flat), flat
    assert flat.index("Invoices 3") > flat.index("Currency EUR"), flat

    # And the grand total says it is the grand total. Without the heading the
    # sheet shows the same labels twice at two scopes and explains neither --
    # a subtotal that equals the total is exactly when that misleads.
    heading = flat.index("Report total")
    assert heading > flat.index("Sales total 4,610.00"), flat
    assert any("Sales total" in text for text in flat[heading:]), flat


def test_a_record_report_puts_its_own_fields_on_sheet_two(
    reporting: tuple[ReportService, RequestContext],
) -> None:
    service, context = reporting
    document = service.build_for_record("sales.invoice", 1, context)
    assert document.record_values, "a record report heads itself with its fields"

    book = _sheets(document)
    assert book.sheetnames == [RECORDS_SHEET, DETAILS_SHEET]
    # The lines are the table; the invoice's own fields are not table.
    assert [cell.value for cell in book[RECORDS_SHEET][1]] == [
        column.label for column in document.detail.columns
    ]
    flat = " ".join(
        str(cell.value)
        for row in book[DETAILS_SHEET].iter_rows()
        for cell in row
        if cell.value is not None
    )
    for value in document.record_values:
        assert value.label in flat


def test_a_flat_document_still_writes_one_sheet_of_rows() -> None:
    """The shape browse export already ships must not change."""

    document = ReportDocument(
        report="x",
        title="X",
        application="App",
        generated_at=__import__("datetime").datetime(
            2026, 8, 24, tzinfo=__import__("datetime").timezone.utc
        ),
        header_text=("App - X",),
        record_values=(),
        detail=ReportTable(
            columns=(ReportColumn("n", "Number"), ReportColumn("t", "Total", "right")),
            rows=(
                (ReportCell("INV-1"), ReportCell("10.50", "right")),
                (ReportCell("=SUM(A1:A9)"), ReportCell("20.50", "right")),
            ),
        ),
        footer_values=(),
        page_footer_template="",
        suggested_filename="x",
    )
    typed = ((None, Decimal("10.50")), (None, Decimal("20.50")))

    book = _sheets(document, typed)
    sheet = book[RECORDS_SHEET]

    assert [cell.value for cell in sheet[1]] == ["Number", "Total"]
    assert sheet.cell(row=2, column=2).value == Decimal("10.50")
    # Still not a formula, whatever it looks like.
    assert sheet.cell(row=3, column=1).value == "=SUM(A1:A9)"
    assert sheet.cell(row=3, column=1).data_type == "s"


def test_a_grouped_document_types_the_columns_it_was_given() -> None:
    """Typed values are positional over the detail rows, groups or not.

    The leading group columns are text -- a group key is a caption -- so the
    typed table lines up with `detail.rows`, not with what the sheet ends up
    showing.
    """

    document = ReportDocument(
        report="x",
        title="X",
        application="App",
        generated_at=__import__("datetime").datetime(
            2026, 8, 24, tzinfo=__import__("datetime").timezone.utc
        ),
        header_text=(),
        record_values=(),
        detail=ReportTable(
            columns=(ReportColumn("n", "Number"), ReportColumn("t", "Total", "right")),
            rows=(
                (ReportCell("INV-1"), ReportCell("10.50", "right")),
                (ReportCell("INV-2"), ReportCell("20.50", "right")),
            ),
        ),
        footer_values=(ReportValue("Sales total", "31.00", "right"),),
        page_footer_template="",
        suggested_filename="x",
        groups=(
            ReportGroup(
                values=(ReportValue("Customer", "ACME"),),
                row_start=0,
                row_count=2,
                footer_values=(ReportValue("Sales total", "31.00", "right"),),
            ),
        ),
    )
    typed = ((None, Decimal("10.50")), (None, Decimal("20.50")))

    sheet = _sheets(document, typed)[RECORDS_SHEET]

    assert [cell.value for cell in sheet[1]] == ["Customer", "Number", "Total"]
    assert sheet.cell(row=2, column=1).value == "ACME"
    # Column 3 on the sheet is column 2 of the detail row.
    assert sheet.cell(row=2, column=3).value == Decimal("10.50")
